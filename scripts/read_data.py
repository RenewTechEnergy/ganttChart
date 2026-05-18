#!/usr/bin/env python3
"""Read tabular data from .xlsx, .xls, .csv, or .tsv into a list of dicts.

Returned shape:
{
    "headers": ["Col A", "Col B", ...],
    "rows": [{"Col A": "val", "Col B": "val", ...}, ...],
    "sheet_names": ["Sheet1", ...],   # for xlsx only; otherwise empty list
    "used_sheet": "Sheet1"            # which one we actually read
}

Dates are returned as ISO strings (YYYY-MM-DD). Everything else is a string.
"""
from __future__ import annotations

import csv
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any


def _iso(value: Any) -> str:
    """Stringify a cell value. Dates become ISO; everything else is str()."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


HEADER_HINTS = (
    "name", "project", "task", "title", "activity",
    "start", "begin", "kickoff", "sched",
    "end", "finish", "due", "deadline", "target",
    "status", "phase", "stage",
    "assigned", "owner", "responsible", "manager",
    "client", "customer", "company",
    "id", "%", "cost", "duration", "predecessor",
)


def _score_header_row(values: list) -> float:
    """Heuristic score for how likely this row is a header row.

    Headers tend to: (a) have many non-empty cells, (b) be mostly short strings,
    (c) contain familiar role keywords, (d) NOT contain dates or many numbers.
    """
    if not values:
        return 0.0
    non_empty = [_iso(v) for v in values if v is not None and _iso(v)]
    if len(non_empty) < 2:
        return 0.0

    # Penalize rows where most cells are dates or pure numbers — those are data
    from datetime import datetime, date
    date_count = sum(1 for v in values if isinstance(v, (datetime, date)))
    num_count  = sum(1 for v in values if isinstance(v, (int, float)) and not isinstance(v, bool))
    if (date_count + num_count) > len(non_empty) / 2:
        return 0.0

    score = float(len(non_empty))

    # Bonus for cells that look like role headers
    lc = " ".join(s.lower() for s in non_empty)
    keyword_hits = sum(1 for h in HEADER_HINTS if h in lc)
    score += keyword_hits * 2.5

    # Penalize very long cells (titles/subtitles tend to be long single sentences)
    avg_len = sum(len(s) for s in non_empty) / len(non_empty)
    if avg_len > 35:
        score *= 0.4

    return score


def _detect_header_row(rows: list[list], scan_limit: int = 8) -> int:
    """Return the 0-indexed row that's most likely to be the header.

    Falls back to 0 if nothing looks compelling.
    """
    best_idx = 0
    best_score = 0.0
    for i, row in enumerate(rows[:scan_limit]):
        s = _score_header_row(row)
        if s > best_score:
            best_score = s
            best_idx = i
    return best_idx


def read_xlsx(path: Path, sheet: str | None = None, header_row: int | None = None) -> dict:
    """Read an xlsx into headers + rows.

    header_row is 1-indexed (matches Excel row numbers). If None, auto-detected.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required to read .xlsx files. "
            "Install with: pip install openpyxl --break-system-packages"
        ) from e

    # First pass: read with data_only=False to detect best sheet (one with most data)
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    used = sheet if sheet and sheet in sheet_names else sheet_names[0]
    ws = wb[used]

    # Slurp the whole sheet so we can detect headers (data_only handles formulas)
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        wb.close()
        return {"headers": [], "rows": [], "sheet_names": sheet_names, "used_sheet": used, "header_row_index": 0}

    if header_row is not None:
        hdr_idx = max(0, header_row - 1)
    else:
        hdr_idx = _detect_header_row(all_rows)

    header_vals = all_rows[hdr_idx]
    headers = [_iso(h) or f"Column {i+1}" for i, h in enumerate(header_vals)]

    rows: list[dict] = []
    for raw in all_rows[hdr_idx + 1:]:
        if raw is None:
            continue
        if all(v is None or _iso(v) == "" for v in raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            row[h] = _iso(raw[i]) if i < len(raw) else ""
        rows.append(row)

    wb.close()
    return {
        "headers": headers,
        "rows": rows,
        "sheet_names": sheet_names,
        "used_sheet": used,
        "header_row_index": hdr_idx,  # 0-indexed; data starts at hdr_idx + 1
    }


def _pick_data_sheet(wb, sheet_names: list[str]) -> str:
    """Pick the sheet with the most rows of actual data — usually the right one."""
    best = sheet_names[0]
    best_count = 0
    for name in sheet_names:
        ws = wb[name]
        # Count rows with at least 3 non-empty cells (likely tabular data)
        count = 0
        for raw in ws.iter_rows(values_only=True):
            non_empty = sum(1 for v in raw if v is not None and _iso(v))
            if non_empty >= 3:
                count += 1
            if count > best_count:
                best_count = count
                best = name
                break  # don't scan all rows; first match wins for this sheet
    return best


def read_csv(path: Path, delim: str | None = None) -> dict:
    text = path.read_text(encoding="utf-8-sig")  # strip BOM
    # Sniff delimiter if not given
    if delim is None:
        first_line = text.split("\n", 1)[0]
        delim = "\t" if first_line.count("\t") > first_line.count(",") else ","

    reader = csv.reader(text.splitlines(), delimiter=delim)
    try:
        header_row = next(reader)
    except StopIteration:
        return {"headers": [], "rows": [], "sheet_names": [], "used_sheet": ""}

    headers = [h.strip() or f"Column {i+1}" for i, h in enumerate(header_row)]

    rows: list[dict] = []
    for raw in reader:
        if not raw or all(not (c or "").strip() for c in raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            row[h] = (raw[i] if i < len(raw) else "").strip()
        rows.append(row)

    return {
        "headers": headers,
        "rows": rows,
        "sheet_names": [],
        "used_sheet": "",
        "header_row_index": 0,
    }


def read_any(path: str | Path, sheet: str | None = None, header_row: int | None = None) -> dict:
    p = Path(path).expanduser().resolve()
    if not p.exists():
        raise FileNotFoundError(f"No such file: {p}")
    suffix = p.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return read_xlsx(p, sheet=sheet, header_row=header_row)
    if suffix in (".csv",):
        return read_csv(p, delim=",")
    if suffix in (".tsv", ".txt"):
        return read_csv(p, delim="\t")
    raise ValueError(f"Unsupported file extension: {suffix}")


if __name__ == "__main__":
    # Quick CLI: python read_data.py <path> [sheet]
    if len(sys.argv) < 2:
        print("Usage: read_data.py <path> [sheet]", file=sys.stderr)
        raise SystemExit(1)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    data = read_any(path, sheet=sheet)
    print(json.dumps(data, indent=2, default=str))
